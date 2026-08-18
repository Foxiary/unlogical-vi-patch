# -*- coding: utf-8 -*-
"""Pull a specific set of cells down from a new sheet snapshot, three-way and guarded.

For the common round: "I fixed a term on the sheet, here is the export."  Not a full
merge — `--match` narrows it to the cells you actually mean, so a terminology pass
does not drag every other unrelated edit along with it.

Three-way, because the user edits the packed game too (see the merge memory):

    new == build                 -> đã có, bỏ qua
    new != build, base == build  -> áp
    new != build, base != build  -> hai bên đều đổi, BÁO rồi bỏ qua

Ids come from the sheet's own ID column:

    76/txt/0011                      -> ScenarioData scenarioID 76, text[11]
    TerminalHomeAlertData/alert/id71 -> json bundle, that asset, field, entry id

Guards on every cell before it is written:

- `[...]` tag multiset must match, so a cell cannot rewrite a lookup key
- `[主人公]` must be present on both sides or neither
- quotes must balance (`"` even, 「」/『』 counts equal)
- **hard line breaks are carried from the build**, never taken from the sheet: the
  `sd_*` Vietnamese column is one flat line, so writing it verbatim flattens the
  layout (1 530 breaks were destroyed that way once).  difflib maps the old break
  positions onto the new wording.

    python tools\\apply_sheet_cells.py --match mainframe
    python tools\\apply_sheet_cells.py --match mainframe --apply
    python tools\\apply_sheet_cells.py --new "…(23).xlsx" --base "…(22).xlsx" --match X
    python tools\apply_sheet_cells.py --take-sheet=71/txt/0064,80/txt/0170 --apply
"""
import difflib
import glob
import io
import json
import os
import re
import shutil
import sys

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")
HERE = os.path.dirname(os.path.abspath(__file__))
ROOT = os.path.dirname(HERE)
sys.path.insert(0, HERE)

import UnityPy   # noqa: E402
from openpyxl import load_workbook   # noqa: E402

SCENARIO = os.path.join(ROOT, "romfs", "Data", "StreamingAssets", "scenario", "scenario01")
JSONB = os.path.join(ROOT, "romfs", "Data", "StreamingAssets", "json", "json")
SNAPSHOTS = r"D:\Downloads\UNLOGICAL_v2*.xlsx"

APPLY = "--apply" in sys.argv


def arg(name, default=None):
    for a in sys.argv:
        if a.startswith("--%s=" % name):
            return a.split("=", 1)[1]
    return default


MATCH = arg("match")
# `--take-sheet=id,id,…`: những ô đã xem bằng mắt và chốt "sheet thắng", dù build cũng đã
# đổi. Sinh ra khi snapshot đã merge bị **tải đè lên cùng tên file** nên không còn bản nền
# nào khớp build: vòng (32) export lại 18/08 làm 27 ô báo "cả hai bên đổi" mà 26 trong số
# đó hai bên chỉ khác dấu nháy. Phải liệt kê id tường minh — không có chế độ "lấy tất".
TAKE = {x.strip() for x in (arg("take-sheet") or "").split(",") if x.strip()}
TAG = re.compile(r"\[[^\[\]\n]*\]")
SD_ID = re.compile(r"^(\d+)/txt/(\d+)$")
DATA_ID = re.compile(r"^([A-Za-z&]+Data)/([A-Za-z_]+)/id(\d+)$")


def backup_path(base):
    """Không bao giờ **bỏ qua** backup vì tên đã tồn tại.

    Tên backup lấy từ tên file snapshot, mà người dùng tải lại **đè lên cùng tên** (memory
    merge: "N là thứ tự tải, không phải thời gian"). Nên `_backup\\scenario01.UNLOGICAL_v2(32)`
    đã có nghĩa là *vòng trước cùng tên sheet*, không phải vòng này — bỏ qua là mất đúng
    cái mốc để lùi một bước. Thật: `(32)` được export lại 18/08 14:11 sau khi vòng `(32)`
    đầu đã merge xong, hai nội dung khác nhau 670 ô.
    """
    if not os.path.exists(base):
        return base
    i = 2
    while os.path.exists("%s-%d" % (base, i)):
        i += 1
    return "%s-%d" % (base, i)


def newest_two():
    cands = sorted(glob.glob(SNAPSHOTS), key=os.path.getmtime, reverse=True)
    if len(cands) < 2:
        raise SystemExit("cần ít nhất hai snapshot khớp " + SNAPSHOTS)
    return cands[0], cands[1]


def read_sheet(path):
    """{id: (tiếng Việt, tiếng Nhật)} — sd_* dịch ở cột D, Nhật ở cột C; tab *Data
    dịch ở cột C, Nhật ở cột B.

    **Làm phẳng `\\n` của sheet ngay ở đây.** Quy ước của cả tool là "sheet giữ câu chữ,
    build giữ ngắt dòng" — nên mọi so sánh phía dưới đều làm phẳng phía build
    (`cur.replace("\\n", " ")`). 56/41.247 ô của snapshot (32) lại có `\\n` thật, và
    những ô đó thì so nguyên văn *không bao giờ* khớp: `80/txt/0166` bị báo "cả hai bên
    đổi" trong khi thực tế chỉ sheet đổi (dấu `‘ ’` cong → `"`). `carry_breaks()` cũng
    giả định phía mới là một dòng phẳng, nên để `\\n` sống sót tới đó là chồng ngắt dòng.
    """
    wb = load_workbook(path, read_only=True, data_only=True)
    out = {}
    for ws in wb.worksheets:
        vcol, jcol = (3, 2) if ws.title.startswith("sd_") else (2, 1)
        for row in ws.iter_rows(values_only=True):
            if not row or not isinstance(row[0], str):
                continue
            key = row[0].strip()
            if not (SD_ID.match(key) or DATA_ID.match(key)):
                continue
            val = row[vcol] if len(row) > vcol and isinstance(row[vcol], str) else ""
            jp = row[jcol] if len(row) > jcol and isinstance(row[jcol], str) else ""
            val = val.replace("\r\n", "\n")
            if ws.title.startswith("sd_"):
                # `sd_*` giữ quy ước cũ: build sở hữu ngắt dòng, sheet chỉ giữ câu chữ.
                val = val.replace("\n", " ")
            else:
                # Tab *Data thì SHEET sở hữu ngắt dòng. Đổi Alt+Enter thành dấu `\n` văn
                # bản để đi tiếp bằng đúng đường của `expand_breaks()`; gộp khoảng trắng
                # hai bên, vì `space + Alt+Enter` mà để nguyên sẽ thành space đôi rồi
                # `carry_breaks` chỉ ăn một cái — đúng lỗi space rác của vòng (40).
                val = re.sub(r"[ \t　]*\n[ \t　]*", BS_N, val)
            out[key] = (val, jp)
    wb.close()
    return out


# Nhãn field trên sheet KHÔNG phải tên field thật, và khoá gốc cũng không luôn là `data`.
# Đo 18/08 trên snapshot (37): trong 11 nhãn sheet dùng chỉ `TerminalHomeAlertData/alert`
# là khớp thẳng, `rule_body` có nhánh riêng, còn lại 262 hàng **không bao giờ áp được** —
# 5 nhãn sai tên field, 4 nhãn sai vì `dj.get("data", dj)` không ra list có `id`.
# Lớp lỗi này ẩn được lâu vì tool chạy theo diff: một nhãn sai chỉ lộ ra ở vòng nào
# đúng mấy hàng đó tình cờ đổi (vòng (39): 6/44 hàng note đổi -> 6 dòng "không có field").
FIELD_MAP = {
    ("GenebarkNoteData", "note"): ("data", "text"),
    ("GenebarkNewsData", "news_title"): ("data", "title"),
    ("GenebarkNewsData", "news_body"): ("data", "text"),
    ("TerminalControlSkillData", "skill_name"): ("data", "request"),
    ("TerminalControlSkillData", "skill_desc"): ("data", "caption"),
    ("TerminalProfileData", "prof_name"): ("info", "name"),
    ("TerminalProfileData", "prof_comment"): ("info", "comment"),
    ("ShortStoryData", "ss_title"): ("list", "title"),
    ("TerminalRuleData", "rule_title"): ("data.items", "title"),
    ("TerminalHomeAlertData", "alert"): ("data", "alert"),
}

# `\n` dạng VĂN BẢN (U+005C U+006E) là cách các tab *Data khai báo ngắt dòng — asset thì
# chỉ dùng U+000A. Snapshot (37) có 2 dấu như vậy ở note id1, khớp đúng 3 dòng của build;
# (39) xoá hết, làm 6 note thành một dòng phẳng. Xem [[unlogical-text-overflow]].
BS_N = chr(92) + "n"        # viết bằng chr() để khỏi lẫn với escape thật


def expand_breaks(s):
    """`\\n` văn bản -> U+000A, GỘP khoảng trắng hai bên.

    Sheet hay gõ `ngày 9 \\n・ Tài liệu` (có space trước dấu), để nguyên thì dòng trên
    mang space đuôi — cùng lớp lỗi với `space + Alt+Enter` thành space đôi.
    """
    return re.sub(r"[ \t　]*%s[ \t　]*" % re.escape(BS_N), "\n", s)


def entries(dj, root):
    """List các mục có `id`, theo đúng khoá gốc của từng asset."""
    if root == "data.items":
        out = []
        for g in dj.get("data", []):
            out.extend(g.get("items", []))
        return out
    rows = dj.get(root)
    return rows if isinstance(rows, list) else []


def field_get(ent, field):
    """Giá trị field, gỡ một tầng `{"jp": …}` nếu có (note/news bọc thế)."""
    v = ent.get(field)
    return v.get("jp", "") if isinstance(v, dict) else v


RULE_ID = re.compile(r"^([A-Za-z&]+Data)/rule_body/id(\d+)$")


def read_rule_rows(path):
    """{(id, trang): tiếng Việt} cho tab TerminalRuleData.

    **Giữ thứ tự hàng**: mỗi id lặp lại một hàng cho mỗi trang, hàng thứ k là
    `content[k]`. Nhét vào dict theo id là gộp 39 hàng thành 21 và tab trông như
    không map được (memory `unlogical-sheet-merge`).
    """
    wb = load_workbook(path, read_only=True, data_only=True)
    out, seen = {}, {}
    for ws in wb.worksheets:
        if not ws.title.startswith("TerminalRuleData"):
            continue
        for row in ws.iter_rows(values_only=True):
            if not row or not isinstance(row[0], str):
                continue
            m = RULE_ID.match(row[0].strip())
            if not m:
                continue
            i = int(m.group(2))
            k = seen.get(i, -1) + 1
            seen[i] = k
            out[(i, k)] = row[2] if len(row) > 2 and isinstance(row[2], str) else ""
    wb.close()
    return out


def merge_rule_text(build, sheet):
    """Lấy CÂU CHỮ của sheet, giữ KHOẢNG TRẮNG ĐẦU DÒNG của build.

    Sheet đã rã hết `　` thành một space ASCII và biến dòng trắng thành một dấu
    cách, nên áp nguyên văn là ép thụt lề còn 1/3 và phá cả bậc bullet của trang
    RULE. None nếu số dòng hai bên khác nhau — chỗ đó cần người xem.
    """
    b, s = build.split("\n"), sheet.split("\n")
    if len(b) != len(s):
        return None
    out = []
    for bl, sl in zip(b, s):
        if not bl.strip():                 # dòng trắng của build giữ nguyên
            out.append(bl)
            continue
        lead = bl[:len(bl) - len(bl.lstrip("　 "))]
        out.append(lead + sl.strip())
    return "\n".join(out)


def duplicate_paste(todo):
    """Bắt lỗi "bản dịch rơi vào sai ô": hai ô đổi trong cùng vòng mà bản dịch mới
    giống nhau từng chữ trong khi bản Nhật của chúng khác nhau — dấu hiệu dán đè.

    Đã bắt được thật ở snapshot (24): `71/txt/0379` mang bản dịch của `71/txt/0377`.
    """
    by_val = {}
    for key, bv, nv, jp in todo:
        by_val.setdefault(nv, []).append((key, jp, bv))
    bad = set()
    for nv, rows in by_val.items():
        if len(rows) < 2 or len({jp for _, jp, _ in rows}) < 2:
            continue
        # Ô nào bị dán đè thì bản mới của nó KHÁC XA bản cũ của chính nó; ô lành chỉ
        # sửa nhẹ. Nhờ vậy không chặn oan ô lành trong cùng nhóm.
        ratio = {key: difflib.SequenceMatcher(None, bv, nv).ratio() for key, _, bv in rows}
        keep = max(ratio, key=ratio.get)
        for key in ratio:
            if key != keep or ratio[keep] < 0.6:
                bad.add(key)
    return bad


def load_text(path, name):
    env = UnityPy.load(path)
    for o in env.objects:
        if o.type.name == "TextAsset" and o.read().m_Name == name:
            d = o.read()
            raw = d.m_Script
            if not isinstance(raw, str):
                raw = bytes(raw).decode("utf-8")
            return env, d, raw
    raise SystemExit("không thấy %s trong %s" % (name, path))


def carry_breaks(old, new):
    """Đặt lại các `\\n` của bản build lên câu chữ mới, theo vị trí ký tự đã khớp."""
    if "\n" not in old:
        return new
    flat = old.replace("\n", " ")
    sm = difflib.SequenceMatcher(None, flat, new, autojunk=False)
    mapping = {}
    for a, b, n in sm.get_matching_blocks():
        for i in range(n):
            mapping[a + i] = b + i
    out = list(new)
    marks = []
    for i, c in enumerate(flat):
        if c == " " and old[i] == "\n":
            j = mapping.get(i)
            if j is not None and j < len(out) and out[j] == " ":
                marks.append(j)
    for j in marks:
        out[j] = "\n"
    got = "".join(out)
    if got.count("\n") != old.count("\n"):
        return None            # không đặt lại đủ -> để người xem quyết
    return got


RUBY = re.compile(r"\[([^\[\]\n']*?)'([^\[\]\n]*?)\]")
DIC = re.compile(r"\[dic\s+no=(\d+)\s+text=([^\[\]\n]*)\]")


def tag_key(t):
    """Khoá so sánh tag: phần nào là KHOÁ TRA CỨU thì so nguyên văn, phần nào chỉ để
    hiển thị thì bỏ qua.

    - `[gốc'ruby]`: cả hai nửa đều là chữ hiển thị, nên **cho phép đảo** (so theo tập
      hợp) — đó là đợt sửa `[Người lựa chọn'Selector]` -> `[Selector'Người lựa chọn]`.
      Đổi *nội dung* một nửa thì vẫn bị chặn.
    - `[dic no=N text=X]`: `no` là khoá, `text` là chữ hiển thị.
    - còn lại (lệnh diễn xuất, `[se file=…]`, `[主人公]`) so nguyên văn.
    """
    m = RUBY.fullmatch(t)
    if m and not t.startswith("[dic "):
        return "RUBY:" + "|".join(sorted([m.group(1), m.group(2)]))
    m = DIC.fullmatch(t)
    if m:
        return "DIC:" + m.group(1)
    return t


def is_ruby(t):
    return bool(RUBY.fullmatch(t)) and not t.startswith("[dic ")


def ruby_change_ok(old, new):
    """Tag ruby được phép đổi theo ba cách: giữ nguyên, **đảo hai nửa**, hoặc **bỏ tag
    mà giữ lại một nửa làm chữ thường** (vòng (28) bỏ gloss: `[Thiên thần tập sự'Spirit]`
    -> `Spirit`). Mất tag mà cả hai nửa cũng mất thì là xoá hụt — chặn."""
    for t in TAG.findall(old):
        if not is_ruby(t):
            continue
        if t in new:
            continue
        m = RUBY.fullmatch(t)
        if "[%s'%s]" % (m.group(2), m.group(1)) in new:
            continue
        if m.group(1) in new or m.group(2) in new:
            continue
        return "mất tag ruby %s mà không giữ lại nửa nào" % t
    return None


def guards(old, new):
    bad = []
    # Ô trắng không bao giờ được ghi đè lên chữ đang có: đúng lớp lỗi đã làm
    # `107/txt/0099` và `0302` rỗng hẳn trên máy (bản Nhật là 「…………」). Tool này so
    # theo diff nên không chạm vào ô không đổi, nhưng một pass "áp nguyên sheet" thì có.
    if not new.strip() and old.strip():
        bad.append("ô sheet trắng mà build đang có chữ")
    # Khoá tra cứu: lệnh diễn xuất, [主人公], [se file=…]… phải khớp từng cái.
    lookup = lambda s: sorted(t for t in TAG.findall(s)                      # noqa: E731
                              if not is_ruby(t) and not DIC.fullmatch(t))
    if lookup(old) != lookup(new):
        bad.append("tag khoá tra cứu bị đổi")
    # Link từ điển: `no=` không được đổi/mất (chữ hiển thị thì tuỳ).
    dic_nos = lambda s: sorted(m.group(1) for m in DIC.finditer(s))          # noqa: E731
    if dic_nos(old) != dic_nos(new):
        bad.append("link [dic no=…] bị đổi hoặc mất")
    r = ruby_change_ok(old, new)
    if r:
        bad.append(r)
    if ("[主人公]" in old) != ("[主人公]" in new):
        bad.append("token [主人公] chỉ có ở một bên")
    if new.count('"') % 2:
        bad.append("số dấu \" lẻ")
    for a, b in (("「", "」"), ("『", "』")):
        if new.count(a) != new.count(b):
            bad.append("ngoặc %s%s không cân" % (a, b))
    return bad


def main():
    new_f, base_f = arg("new"), arg("base")
    if not new_f or not base_f:
        n, b = newest_two()
        new_f, base_f = new_f or n, base_f or b
    print("sheet mới : %s" % new_f)
    print("sheet nền : %s" % base_f)
    print("lọc       : %s" % (MATCH or "(không lọc — mọi ô đã đổi)"))
    new, base = read_sheet(new_f), read_sheet(base_f)

    pat = re.compile(MATCH, re.I) if MATCH else None
    todo = []
    for key, (nv, jp) in new.items():
        b = base.get(key)
        if b is None or nv == b[0]:
            continue
        if pat and not (pat.search(nv) or pat.search(b[0])):
            continue
        todo.append((key, b[0], nv, jp))
    print("ô đã đổi trên sheet và khớp bộ lọc: %d" % len(todo))
    dup = duplicate_paste(todo)
    if dup:
        print("\n!! %d ô có DẤU HIỆU DÁN ĐÈ trên sheet "
              "(bản dịch mới trùng nhau mà bản Nhật khác nhau)" % len(dup))
        for key in sorted(dup):
            jp = next(j for k, _, _, j in todo if k == key)
            print("   %-14s JP: %r" % (key, jp[:64].replace("\n", "⏎")))
        print("   -> bỏ qua những ô này; sửa trên sheet rồi chạy lại")
        todo = [t for t in todo if t[0] not in dup]
    print()
    if not todo:
        return

    env_s, d_s, raw_s = load_text(SCENARIO, "ScenarioData")
    data_s = json.loads(raw_s.lstrip("﻿"))
    sid_map = {t["scenarioID"]: ti for ti, t in enumerate(data_s["target"])}
    out_s, changed_s = raw_s, []
    json_edits = []

    for key, bv, nv, _jp in sorted(todo):
        m = SD_ID.match(key)
        if m:
            sid, idx = int(m.group(1)), int(m.group(2))
            ti = sid_map.get(sid)
            if ti is None:
                print("! %s: không có scenarioID %d" % (key, sid)); continue
            cur = data_s["target"][ti]["text"][idx]
            kind = "ScenarioData"
        elif RULE_ID.match(key):
            continue                     # xử lý riêng ở nhánh rule_body bên dưới
        else:
            m = DATA_ID.match(key)
            if not m:
                print("! %s: id lạ" % key); continue
            asset, sfield, eid = m.group(1), m.group(2), int(m.group(3))
            root, field = FIELD_MAP.get((asset, sfield), ("data", sfield))
            _, _, raw_j = load_text(JSONB, asset)
            dj = json.loads(raw_j.lstrip("﻿"))
            ent = next((e for e in entries(dj, root) if e.get("id") == eid), None)
            if ent is None:
                print("! %s: không có id %d trong %s (khoá gốc %r)"
                      % (key, eid, asset, root)); continue
            # Tên field của sheet phải giải được ra field thật. Đọc `""` cho field không
            # tồn tại thì ô trông như "build trống" và bị báo sai là "cả hai bên đổi".
            if field not in ent:
                print("! %s: %s không có field %r (nhãn sheet %r; có: %s)"
                      % (key, asset, field, sfield, ", ".join(sorted(ent)))); continue
            cur = field_get(ent, field)
            kind = asset

        # So ba chiều trên bản ĐÃ LÀM PHẲNG: build giữ `\n` mà sheet thì không, nên
        # so nguyên văn sẽ báo "cả hai bên đổi" cho cả những ô build vốn đã đúng.
        # Tab *Data khai ngắt dòng bằng `\n` VĂN BẢN, asset thì dùng U+000A. Phải mở ra
        # trước khi so, không thì ô nào có dấu đó cũng "không bao giờ khớp".
        nv_x = expand_breaks(nv)                     # bản có ngắt dòng do SHEET khai
        nv, bv = nv_x.replace("\n", " "), expand_breaks(bv).replace("\n", " ")
        flat_cur = cur.replace("\n", " ")
        # "Đã có bản mới" phải so NGUYÊN VĂN khi sheet tự khai ngắt dòng — không thì ô chỉ
        # khác bố cục mà giống câu chữ sẽ bị bỏ qua, và cấu trúc mới không bao giờ xuống
        # (vòng (40): 4 hàng note đúng kiểu đó).
        if (cur == nv_x) if "\n" in nv_x else (flat_cur == nv):
            print("=  %-34s đã có bản mới" % key); continue
        if flat_cur != bv and key in TAKE:
            # Người đã xem và chốt "sheet thắng ô này". Vẫn qua đủ các chốt còn lại
            # (tag khoá tra cứu, `no=` của [dic], [主人公], ngoặc cân, carry_breaks) —
            # `--take-sheet` chỉ bỏ *một* điều kiện: "build chưa ai sửa".
            print("~  %-34s LẤY THEO SHEET (--take-sheet, đè bản build)" % key)
            print("      build bị đè: %r" % flat_cur[:88])
        elif flat_cur != bv:
            print("!! %-34s CẢ HAI BÊN ĐỔI — bỏ qua" % key)
            print("      nền   : %r" % bv[:88])
            print("      build : %r" % flat_cur[:88])
            print("      sheet : %r" % nv[:88])
            continue
        bad = guards(cur, nv)
        if bad:
            print("!! %-34s chốt chặn: %s" % (key, "; ".join(bad))); continue
        # Sheet có tự khai ngắt dòng thì SHEET thắng — chỉ các tab *Data làm được, bằng
        # `\n` văn bản. Không khai thì đắp lại ngắt dòng của build như cũ.
        if "\n" in nv_x:
            val = nv_x
        else:
            val = carry_breaks(cur, nv)
            if val is None:
                print("!! %-34s không đặt lại được %d ngắt dòng — bỏ qua"
                      % (key, cur.count("\n"))); continue
        # Space ASCII đứng ngay trước một ngắt dòng cứng thì không bao giờ có nghĩa, mà
        # `carry_breaks` sinh ra nó khi ô sheet có space đôi: `71/txt/0344` ra `. ⏎Cô`.
        # (Chỉ cắt space/tab — `　` SAU ngắt dòng là thụt lề thật, phải giữ.)
        val = re.sub(r"[ \t]+\n", "\n", val)
        # Chốt chặn LÀM PHẲNG: không ô nào được mất ngắt dòng so với build. Đây đúng là
        # lớp hỏng đã phá 1 530 ngắt dòng một lần, và vòng (39) tái diễn — 6 hàng note bị
        # xoá hết dấu `\n` văn bản nên sẽ làm phẳng note 3 dòng thành 1.
        # …nhưng chỉ khi sheet KHÔNG tự khai bố cục. Sheet có khai thì nó là bên có thẩm
        # quyền và việc giảm dòng là chủ ý: vòng (40) id6 nối lại đúng 3 dòng như bản gốc
        # (build đang 4 dòng, ngắt giữa cụm "cách nào / khác là"), chặn là chặn oan.
        if "\n" not in nv_x and val.count("\n") < cur.count("\n"):
            print("!! %-34s chốt chặn: sheet làm phẳng, mất %d ngắt dòng (build %d -> %d)"
                  % (key, cur.count("\n") - val.count("\n"),
                     cur.count("\n"), val.count("\n"))); continue

        print("-> %-34s %s" % (key, kind))
        print("      cũ : %r" % cur[:88].replace("\n", "⏎"))
        print("      mới: %r" % val[:88].replace("\n", "⏎"))
        if kind == "ScenarioData":
            changed_s.append((ti, sid, idx, cur, val))
        else:
            json_edits.append((kind, field, eid, cur, val, root))

    # ---- ScenarioData: text[] + bản sao scriptText
    # Vá theo CẢ MẢNG `text[]` của từng target, không theo từng chuỗi: có những câu
    # trùng nhau từng chữ (70/txt/1216 và 1218), thay theo chuỗi sẽ đụng cả hai.
    if changed_s:
        def enc(x):
            return json.dumps(x, ensure_ascii=False, separators=(",", ":"))
        for ti in sorted({c[0] for c in changed_s}):
            arr_old = list(data_s["target"][ti]["text"])
            arr_new = list(arr_old)
            for t2, sid, idx, cur, val in [c for c in changed_s if c[0] == ti]:
                assert arr_new[idx] == cur, "text[%d] không như đã đọc" % idx
                arr_new[idx] = val
            oj, nj = enc(arr_old), enc(arr_new)
            if out_s.count(oj) != 1:
                raise SystemExit("mảng text[] của target[%d] khớp %d lần" % (ti, out_s.count(oj)))
            out_s = out_s.replace(oj, nj)
        for ti in sorted({c[0] for c in changed_s}):
            script = cur_script = data_s["target"][ti]["scriptText"]
            for t2, sid, idx, cur, val in [c for c in changed_s if c[0] == ti]:
                lines = cur_script.split("\n")
                ol = cur.split("\n")
                hits = [k for k in range(len(lines) - len(ol) + 1) if lines[k:k + len(ol)] == ol]
                if len(hits) == 1:
                    lines[hits[0]:hits[0] + len(ol)] = val.split("\n")
                    cur_script = "\n".join(lines)
                else:
                    print("   (scriptText sID=%s text[%d]: khớp %d lần, bỏ qua mirror)"
                          % (sid, idx, len(hits)))
            if cur_script != script:
                oj, nj = (json.dumps(script, ensure_ascii=False),
                          json.dumps(cur_script, ensure_ascii=False))
                if out_s.count(oj) != 1:
                    raise SystemExit("scriptText target[%d] khớp %d lần" % (ti, out_s.count(oj)))
                out_s = out_s.replace(oj, nj)

    # ---- nhánh rule_body: map theo (id, trang), giữ khoảng trắng đầu dòng của build
    rule_edits = []
    rn, rb = read_rule_rows(new_f), read_rule_rows(base_f)
    rkeys = [k for k in rn if k in rb and rn[k] != rb[k]]
    if rkeys:
        _, _, raw_r = load_text(JSONB, "TerminalRuleData")
        tr = json.loads(raw_r.lstrip("﻿"))
        items = {it["id"]: it for g in tr["data"] for it in g["items"]}
        print("hàng rule_body đổi trên sheet: %d" % len(rkeys))
        for (i, k) in sorted(rkeys):
            it = items.get(i)
            if it is None or k >= len(it.get("content", [])):
                print("!! rule_body id%d trang %d: không có trong asset" % (i, k))
                continue
            cur = it["content"][k]["text"]
            val = merge_rule_text(cur, rn[(i, k)])
            if val is None:
                print("!! rule_body id%d trang %d: số dòng lệch (build %d / sheet %d) — bỏ qua"
                      % (i, k, cur.count("\n") + 1, rn[(i, k)].count("\n") + 1))
                continue
            if val == cur:
                print("=  rule_body id%d trang %d: đã có bản mới" % (i, k))
                continue
            # So ba chiều như nhánh ScenarioData. Nhánh này vốn thiếu chốt đó, nên một
            # bản sửa làm thẳng trên build bị vòng merge sau **âm thầm lật lại** — đã
            # xảy ra thật: `fix_terminal_term.py` đổi id30/44/45 thành "Terminal", sheet
            # vẫn ghi "terminal", vòng chạy lại sẽ hạ chữ hoa xuống.
            # Dùng chính `merge_rule_text` với sheet NỀN: ra đúng `cur` thì build chưa ai
            # sửa; khác `cur` thì hai bên đều đổi.
            base_val = merge_rule_text(cur, rb[(i, k)])
            if base_val is not None and base_val != cur:
                print("!! rule_body id%d trang %d: CẢ HAI BÊN ĐỔI — bỏ qua" % (i, k))
                for l in difflib.unified_diff(base_val.split("\n"), cur.split("\n"),
                                              "nền", "build", lineterm="", n=0):
                    if l[:1] in "+-" and l[:3] not in ("+++", "---"):
                        print("      %s" % l[:96])
                continue
            if val.count("　") < cur.count("　"):
                print("!! rule_body id%d trang %d: mất %d thụt lề `　` — bỏ qua"
                      % (i, k, cur.count("　") - val.count("　")))
                continue
            bad = guards(cur, val)
            if bad:
                print("!! rule_body id%d trang %d: %s" % (i, k, "; ".join(bad)))
                continue
            diff = [l for l in difflib.unified_diff(cur.split("\n"), val.split("\n"),
                                                    lineterm="", n=0)
                    if l[:1] in "+-" and l[:3] not in ("+++", "---")]
            print("-> rule_body id%d trang %d  (%d dòng đổi)" % (i, k, len(diff) // 2))
            for l in diff[:4]:
                print("      %s" % l[:96])
            rule_edits.append((i, k, cur, val))

    print("\náp được: %d ô ScenarioData, %d ô bundle json, %d trang rule_body"
          % (len(changed_s), len(json_edits), len(rule_edits)))
    if not APPLY:
        print("\nCHẠY THỬ — thêm --apply để ghi")
        return

    tag = os.path.splitext(os.path.basename(new_f))[0].replace(" ", "")
    json_bak = [None]        # bundle json bị hai nhánh ghi, chỉ chụp backup một lần
    if changed_s:
        bak = backup_path(os.path.join(ROOT, "_backup", "scenario01.%s" % tag))
        shutil.copy2(SCENARIO, bak); print("backup ->", bak)
        d_s.m_Script = ("﻿" if raw_s.startswith("﻿") else "") + out_s.lstrip("﻿")
        d_s.save()
        with open(SCENARIO, "wb") as f:
            f.write(env_s.file.save(packer="lz4"))
        print("đã ghi", SCENARIO, os.path.getsize(SCENARIO))
        _, _, back = load_text(SCENARIO, "ScenarioData")
        rd = json.loads(back.lstrip("﻿"))
        for ti, sid, idx, cur, val in changed_s:
            assert rd["target"][ti]["text"][idx] == val, "đọc lại %s text[%d] sai" % (sid, idx)
            assert rd["target"][ti]["loadLine"] == data_s["target"][ti]["loadLine"]
        print("  đọc lại: %d ô khớp, loadLine nguyên vẹn" % len(changed_s))

    if rule_edits:
        if json_bak[0] is None:
            json_bak[0] = backup_path(os.path.join(ROOT, "_backup", "json.%s" % tag))
            shutil.copy2(JSONB, json_bak[0]); print("backup ->", json_bak[0])
        env_r, d_r, raw_r = load_text(JSONB, "TerminalRuleData")
        out_r = raw_r
        for i, k, cur, val in rule_edits:
            oj, nj = json.dumps(cur, ensure_ascii=False), json.dumps(val, ensure_ascii=False)
            if out_r.count(oj) != 1:
                raise SystemExit("rule_body id%d trang %d khớp %d lần" % (i, k, out_r.count(oj)))
            out_r = out_r.replace(oj, nj)
        d_r.m_Script = ("﻿" if raw_r.startswith("﻿") else "") + out_r.lstrip("﻿")
        d_r.save()
        with open(JSONB, "wb") as f:
            f.write(env_r.file.save(packer="lz4"))
        print("đã ghi", JSONB, os.path.getsize(JSONB))
        _, _, back = load_text(JSONB, "TerminalRuleData")
        tr2 = json.loads(back.lstrip("﻿"))
        items2 = {it["id"]: it for g in tr2["data"] for it in g["items"]}
        for i, k, cur, val in rule_edits:
            assert items2[i]["content"][k]["text"] == val, \
                "đọc lại rule_body id%d trang %d sai" % (i, k)
        print("  đọc lại: %d trang rule_body khớp" % len(rule_edits))

    if json_edits:
        if json_bak[0] is None:
            json_bak[0] = backup_path(os.path.join(ROOT, "_backup", "json.%s" % tag))
            shutil.copy2(JSONB, json_bak[0]); print("backup ->", json_bak[0])
        for asset in sorted({e[0] for e in json_edits}):
            env_j, d_j, raw_j = load_text(JSONB, asset)
            out_j = raw_j
            for a, field, eid, cur, val, root in [e for e in json_edits if e[0] == asset]:
                oj, nj = (json.dumps(cur, ensure_ascii=False), json.dumps(val, ensure_ascii=False))
                if out_j.count(oj) != 1:
                    raise SystemExit("%s id%d: chuỗi cũ khớp %d lần" % (asset, eid, out_j.count(oj)))
                out_j = out_j.replace(oj, nj)
            d_j.m_Script = ("﻿" if raw_j.startswith("﻿") else "") + out_j.lstrip("﻿")
            d_j.save()
            with open(JSONB, "wb") as f:
                f.write(env_j.file.save(packer="lz4"))
            print("đã ghi", JSONB, os.path.getsize(JSONB))
            _, _, back = load_text(JSONB, asset)
            dj_back = json.loads(back.lstrip("﻿"))
            for a, field, eid, cur, val, root in [e for e in json_edits if e[0] == asset]:
                ent = next(e for e in entries(dj_back, root) if e.get("id") == eid)
                assert field_get(ent, field) == val, "đọc lại %s id%d sai" % (asset, eid)
            print("  đọc lại: %s khớp" % asset)


main()
